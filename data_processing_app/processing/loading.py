import os, csv, io, pandas as pd, msoffcrypto
import xlrd

from openpyxl import load_workbook
from utils.table_utils import pad_rows, make_unique_columns
from config.constants import CSV_SNIFF_BYTES
from workspace.jobs import CANCELLED_MSG

class FileLoader:
    def __init__(self, header_detector, cleaner, logger, password_callback):
        self.headers = header_detector
        self.cleaner = cleaner
        self.logger = logger
        self.password_callback = password_callback
        
    def _check_cancel(self, cancel_event) -> None:
        if cancel_event is not None and cancel_event.is_set():
            raise RuntimeError(CANCELLED_MSG)

    def _choose_text_encoding(self, filename: str):
        with open(filename, "rb") as f:
            sample = f.read(CSV_SNIFF_BYTES)
        candidates = ["utf-8-sig", "utf-8", "cp1252"]
        for enc in candidates:
            try:
                sample.decode(enc, errors="strict")
                return enc, "strict"
            except Exception:
                pass
        return "cp1252", "replace"
    
    def load_file(self, filename, header_cleaning_mode="none", cancel_event=None):
        self.logger.log(f"[LOAD] Loading file: {os.path.basename(filename)}", "green")
        self._check_cancel(cancel_event)


        if filename.lower().endswith((".csv", ".txt", ".f")):
            return self._load_csv(filename, header_cleaning_mode, cancel_event=cancel_event)
        if filename.lower().endswith((".xls", ".xlsx")):
            return self._load_excel(filename, header_cleaning_mode, cancel_event=cancel_event)
        raise ValueError("Unsupported file type")

    def _load_csv(self, filename, header_cleaning_mode="none", cancel_event=None):
        encodings = [
            ("utf-8-sig", "strict"),
            ("utf-8", "strict"),
            ("cp1252", "strict"),
            ("latin-1", "strict"),
        ]
        last_error = None

        for enc, err_mode in encodings:
            try:
                delimiter = self._detect_delimiter(filename, enc, err_mode)
                rows = []
                with open(filename, encoding=enc, errors=err_mode, newline="") as f:
                    rdr = csv.reader(f, delimiter=delimiter)
                    for row in rdr:
                        self._check_cancel(cancel_event)
                        rows.append(row)
                self.logger.log(f"[LOAD] CSV decode: {enc} ({err_mode})", "green")
                return self._process_rows(
                    rows,
                    header_cleaning_mode=header_cleaning_mode,
                    cancel_event=cancel_event,
                )
            except UnicodeDecodeError as e:
                last_error = e
                continue

        # final permissive fallback
        enc, err_mode = self._choose_text_encoding(filename)
        delimiter = self._detect_delimiter(filename, enc, err_mode)
        rows = []
        with open(filename, encoding=enc, errors=err_mode, newline="") as f:
            rdr = csv.reader(f, delimiter=delimiter)
            for row in rdr:
                self._check_cancel(cancel_event)
                rows.append(row)

        self.logger.log(f"[LOAD] CSV decode fallback: {enc} ({err_mode})", "yellow")
        return self._process_rows(
            rows,
            header_cleaning_mode=header_cleaning_mode,
            cancel_event=cancel_event,
        )

    def _load_excel(self, filename, header_cleaning_mode="none", cancel_event=None):
        self._check_cancel(cancel_event)
        ext = os.path.splitext(filename)[1].lower()

        def _sheet_rows_from_workbook(wb):
            ws = None
            try:
                if wb.active is not None and getattr(wb.active, "sheet_state", "visible") == "visible":
                    ws = wb.active
            except Exception:
                ws = None

            if ws is None:
                for candidate in wb.worksheets:
                    if getattr(candidate, "sheet_state", "visible") == "visible":
                        ws = candidate
                        break

            if ws is None:
                raise ValueError("No visible worksheets found in workbook.")

            rows = []
            for row in ws.iter_rows(values_only=True):
                self._check_cancel(cancel_event)
                rows.append(["" if v is None else str(v) for v in row])
            return rows

        normal_exc = None
        try:
            self._check_cancel(cancel_event)

            if ext == ".xlsx":
                wb = load_workbook(filename, read_only=True, data_only=True)
                try:
                    self._check_cancel(cancel_event)
                    rows = _sheet_rows_from_workbook(wb)
                finally:
                    try:
                        wb.close()
                    except Exception:
                        pass

            elif ext == ".xls":
                book = xlrd.open_workbook(filename)
                sheet = book.sheet_by_index(0)
                rows = []
                for r in range(sheet.nrows):
                    self._check_cancel(cancel_event)
                    row = []
                    for c in range(sheet.ncols):
                        v = sheet.cell_value(r, c)
                        row.append("" if v is None else str(v))
                    rows.append(row)

            else:
                raise ValueError("Unsupported Excel file type")

            self._check_cancel(cancel_event)
            df = pd.DataFrame(rows).fillna("")
            return self._process_rows(
                df.values.tolist(),
                header_cleaning_mode=header_cleaning_mode,
                cancel_event=cancel_event,
            )

        except Exception as e:
            normal_exc = e
            if not self.password_callback:
                raise

        if ext != ".xlsx":
            raise normal_exc

        with open(filename, "rb") as f:
            office = msoffcrypto.OfficeFile(f)
            if not office.is_encrypted():
                raise normal_exc

        while True:
            self._check_cancel(cancel_event)

            password = self.password_callback("Enter password:")
            if password is None:
                self.logger.log("[LOAD] Password entry cancelled.", "yellow")
                return None, None

            password = str(password).strip()
            if not password:
                self.logger.log("[LOAD] Empty password entered.", "yellow")
                continue

            try:
                self._check_cancel(cancel_event)

                decrypted = io.BytesIO()
                with open(filename, "rb") as f:
                    office = msoffcrypto.OfficeFile(f)
                    office.load_key(password=password)
                    office.decrypt(decrypted)

                self._check_cancel(cancel_event)
                decrypted.seek(0)

                wb = load_workbook(decrypted, read_only=True, data_only=True)
                try:
                    self._check_cancel(cancel_event)
                    rows = _sheet_rows_from_workbook(wb)
                finally:
                    try:
                        wb.close()
                    except Exception:
                        pass

                self._check_cancel(cancel_event)

                df = pd.DataFrame(rows).fillna("")
                return self._process_rows(
                    df.values.tolist(),
                    header_cleaning_mode=header_cleaning_mode,
                    cancel_event=cancel_event,
                )

            except RuntimeError as e:
                if str(e) == CANCELLED_MSG:
                    raise
                raise
            except Exception:
                self.logger.log("[LOAD] Incorrect password — try again or cancel.", "red")
    
    def _process_rows(self, rows, header_cleaning_mode="none", cancel_event=None):
        self._check_cancel(cancel_event)
        if len(rows) < 4:
            raise ValueError("Invalid file.")
        r1, r2, r3 = rows[:3]
        result = self.headers.detect_header(r1, r2, r3)
        self.headers.last_header_result = result

        has_header, cols, data = self.headers.apply_header_result(result, rows)

        padded, max_cols = pad_rows(data)
        self._check_cancel(cancel_event)

        def _is_blank(v) -> bool:
            return str(v).strip() == ""

        if max_cols > 0 and padded:
            keep_idx = []
            for j in range(max_cols):
                any_nonblank = False
                for row in padded:
                    if j < len(row) and not _is_blank(row[j]):
                        any_nonblank = True
                        break
                if any_nonblank:
                    keep_idx.append(j)

            if len(keep_idx) != max_cols:
                padded = [[row[j] if j < len(row) else "" for j in keep_idx] for row in padded]

                if cols:
                    cols = [cols[j] for j in keep_idx if j < len(cols)]

                max_cols = len(keep_idx)

        if len(cols) < max_cols:
            cols += [f"Column{len(cols) + i + 1}" for i in range(max_cols - len(cols))]

        df = pd.DataFrame(padded, columns=cols).astype(object)
        df = self.cleaner.cleanse_dataframe(df)

        df, has_header = self.headers.analyze_and_log_header(df, has_header)
        self._check_cancel(cancel_event)

        df = self.cleaner.clean_header_names(df, has_header, mode=header_cleaning_mode)

        if has_header:
            new_cols = make_unique_columns(df.columns)
            if list(df.columns) != new_cols:
                self.logger.log("[HEADER] Renamed duplicate columns", "yellow")
                df.columns = new_cols
        return df, has_header

    def _detect_delimiter(self, filename, encoding: str, errors: str):
        with open(filename, encoding=encoding, errors=errors, newline="") as f:
            sample = f.read(CSV_SNIFF_BYTES)

        candidates = [",", ";", "\t", "|"]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters="".join(candidates))
            return dialect.delimiter
        except Exception:
            pass

        counts = {d: sample.count(d) for d in candidates}
        best = max(counts, key=counts.get)
        if counts[best] < 2:
            return ","
        return best